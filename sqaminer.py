#AHparser

import openpyxl

wb = openpyxl.Workbook()
ws= wb.active 

infile = 'h-course-spec-biology.pdf.txt'

part =''
sectnum = 0
sect = ''
subsectnum =0
subsect=''
subsubsectnum =0
subsubsect = ''
lastline =''
LO = ''
currentrow=1
ofh = open('h-biology.tsv', 'w')
with open(infile) as fh:
    line=fh.readline()
    while not line.startswith('===='):
        line=fh.readline().strip()
    
    for line in fh:
        line=line.strip()
        if line.startswith('===='):
            break
        if line.strip() == part:
            continue
        if not line:
            continue
        if line.startswith('#'):
            part=line[1:]
        elif line[0] in '1234567890' and line[1]==' ':
            num,sect = line.strip().split(maxsplit=1)
            sectnum = int(num)
            subsect=''
            subsectnum=''
            subsubsect=''
            subsubsectnum=''
        elif line.startswith('('):
            if line[1] in 'abcdefgh':
                pnum,subsect = line.strip().split(maxsplit=1)
                subsectnum = pnum[1:-1]
                subsubsect=''
                subsubsectnum=''
                ws.append([part,sectnum,sect,subsectnum,subsect,subsubsectnum,subsubsect,line[3:]])
                currentrow +=1
                print("\t".join([str(x) for x in [part,sectnum,sect,subsectnum,subsect,subsubsectnum,subsubsect,line[3:]]]), file=ofh)
            else:
                pnum,subsubsect = line.strip().split(maxsplit=1)
                subsubsectnum = pnum[1:-1]
        elif line.startswith('?'):
            try:
                ws.cell(row=currentrow-1, column=8).value= ws.cell(row=currentrow-1, column=8).value +"\n"+line
            except Exception as e:
                print([x.value for x in ws[currentrow]], currentrow, ws.cell(currentrow-1,8).value,ws.cell(currentrow,8).value,line)
                break
        else:    
            ws.append([part,sectnum,sect,subsectnum,subsect,subsubsectnum,subsubsect,line])
            currentrow +=1
            print("\t".join([str(x) for x in [part,sectnum,sect,subsectnum,subsect,subsubsectnum,subsubsect,line]]), file=ofh)
            if len(line) >3:
                lastline=line.strip() 
ofh.close()
wb.save('h-biology.xlsx')