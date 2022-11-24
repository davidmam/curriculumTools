# -*- coding: utf-8 -*-
"""
Created on Thu Nov 24 10:08:27 2022

@author: dmamartin
"""

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import redis
import os
from redisgraph import Node, Edge, Graph, Path

r = redis.Redis(host='redis-19330.c226.eu-west-1-3.ec2.cloud.redislabs.com', port=19330, username='cdev', password=open('redispassword.txt').read())

redis_graph = Graph('curriculum', r)

wb=openpyxl.open("h-chemistry.xlsx")
ws=wb.active
# create SQA qualification node.
params = {
    'name':'Higher',
    'year':'2022-23',
    'SQCFlevel':6,
    'SQCFcredits':24,
    'subject':'Chemistry'
    }

redis_graph.query('MERGE (:SQAqual {name:$name, year:$year,SQCFlevel:$SQCFlevel, SQCFcredits:$SQCFcredits,subject:$subject})',params)

row = 1
data = [str(x.value ) for x in ws[row]]


while data[7]!='None':
    params = {
        'section_number':data[1],
        'section':data[2],
        'subsection_number':data[3],
        'subsection':data[4],
        'subsubsection_number':data[5],
        'subsubsection':data[6],
        'outcome':data[7],
        'subject':'Chemistry',
        'name':'Higher'
        }
    redis_graph.query('''MATCH (s:SQAqual {name:$name, subject:$subject}) CREATE (s) -[:HAS_OUTCOME]-> (o:SQAoutcome {section_number:$section_number, section:$section, subsection_number:$subsection_number, subsection:$subsection, subsubsection_number:$subsubsection_number, subsubsection:$subsubsection,outcome:$outcome}) ''',params)
    row=row+1
    data = data = [str(x.value ) for x in ws[row]]
    
wb=openpyxl.open("h-biology.xlsx")
ws=wb.active
# create SQA qualification node.
params = {
    'name':'Higher',
    'year':'2022-23',
    'SQCFlevel':6,
    'SQCFcredits':24,
    'subject':'Biology'
    }

redis_graph.query('MERGE (:SQAqual {name:$name, year:$year,SQCFlevel:$SQCFlevel, SQCFcredits:$SQCFcredits,subject:$subject})',params)

row = 1
data = [str(x.value ) for x in ws[row]]


while data[7]!='None':
    params = {
        'topic':data[0],
        'section_number':data[1],
        'section':data[2],
        'subsection_number':data[3],
        'subsection':data[4],
        'subsubsection_number':data[5],
        'subsubsection':data[6],
        'outcome':data[7],
        'subject':'Biology',
        'name':'Higher'
        }
    redis_graph.query('''MATCH (s:SQAqual {name:$name, subject:$subject}) CREATE (s) -[:HAS_OUTCOME]-> (o:SQAoutcome {topic:$topic,section_number:$section_number, section:$section, subsection_number:$subsection_number, subsection:$subsection, subsubsection_number:$subsubsection_number, subsubsection:$subsubsection,outcome:$outcome}) ''',params)
    row=row+1
    data = data = [str(x.value ) for x in ws[row]]

wb=openpyxl.open("ah-biology.xlsx")
ws=wb.active
# create SQA qualification node.
params = {
    'name':'Advanced Higher',
    'year':'2022-23',
    'SQCFlevel':7,
    'SQCFcredits':24,
    'subject':'Biology'
    }

redis_graph.query('MERGE (:SQAqual {name:$name, year:$year,SQCFlevel:$SQCFlevel, SQCFcredits:$SQCFcredits,subject:$subject})',params)

row = 1
data = [str(x.value ) for x in ws[row]]


while data[7]!='None':
    params = {
        'topic':data[0],
        'section_number':data[1],
        'section':data[2],
        'subsection_number':data[3],
        'subsection':data[4],
        'subsubsection_number':data[5],
        'subsubsection':data[6],
        'outcome':data[7],
        'subject':'Biology',
        'name':'Advanced Higher'
        }
    redis_graph.query('''MATCH (s:SQAqual {name:$name, subject:$subject}) CREATE (s) -[:HAS_OUTCOME]-> (o:SQAoutcome {topic:$topic,section_number:$section_number, section:$section, subsection_number:$subsection_number, subsection:$subsection, subsubsection_number:$subsubsection_number, subsubsection:$subsubsection,outcome:$outcome}) ''',params)
    row=row+1
    data = data = [str(x.value ) for x in ws[row]]