# -*- coding: utf-8 -*-
"""
Created on Thu Jun  9 11:29:04 2022

@author: DMAMartin
"""

from tkinter import Tk
from tkinter.filedialog import askdirectory
import os
import openpyxl
from MRS import find_modules2023

def main():
    root = Tk()
    
    TAdir = askdirectory(title="TA folder")
    root.destroy()
    mrs = find_modules2023(TAdir)
    mods = list(mrs.keys())
   
    basemod =''
    for m in mods:
        if m[:7]=="BS31003":
            basemod = m
            break
    studentgrades=extractStudents(mrs[basemod]) # should get students from BS11001
    for mod in mods:
        if mod in mrs and mod[2] in '3':
            print(f'extracting grades from {mod}')
            grades = extractgrades(mrs[mod])
            for s in grades:
                if s in studentgrades:
                    studentgrades[s][mod]=grades[s]
    scores={}
    corescores={}
    credits={}
    #coremoduleset =['BS21001','BS21002','BS21012', 'BS22001', 'BS22002', 'BS22003']
    coremoduleset =['BS31003','BS31004','BS31005', 'BS31006', 'BS32011', 'BS32012']
    for s in studentgrades:
        scores[s]=0
        corescores[s]=0
        credittotal=0
        corecredittotal=0
        for m in studentgrades[s]:
            if m[:2]=='BS':
                credit=20
                if m.split()[0] in ['BS21001', 'BS21002']:
                    credit=10
                if m[2] in '34':
                    credit=15
                try:
                    scores[s] += credit*studentgrades[s][m]
                    if m in coremoduleset:
                        corescores[s] += credit*studentgrades[s][m]
                        corecredittotal += credit
                except:
                    print('could not multiply {} by {}'.format(studentgrades[s][m], credit))
                    scores[s] +=0
                credittotal+=credit
                 
        scores[s]=scores[s]/credittotal
        credits[s]=credittotal
        if corescores[s]:
            corescores[s] =corescores[s]/corecredittotal
    ofh=open("L3 Bio mean scores.txt ", "w")
    for s in studentgrades:
        g = studentgrades[s]
        print(s, g['firstname'], g['lastname'], g['route'], scores[s], corescores[s], credits[s], sep="\t", file=ofh)
    ofh.close()
    
def extractgrades(file):
    workbook=openpyxl.load_workbook(filename=file, data_only=True)
    sheet = workbook['OVERALL Results']
    row = 5
    grades={}
    matric = sheet.cell(row=row, column=1).value
    grade = sheet.cell(row=row, column=8).value
    while matric:
        grades[matric]=grade
        row +=1
        matric = sheet.cell(row=row, column=1).value
        grade = sheet.cell(row=row, column=8).value
    return grades

def extractStudents(file):
    workbook=openpyxl.load_workbook(filename=file, data_only=True)
    sheet = workbook['OVERALL Results']
    row = 5
    students={}
    matric = sheet.cell(row=row, column=1).value
    lastname = sheet.cell(row=row, column=2).value
    firstname = sheet.cell(row=row, column=3).value
    route = sheet.cell(row=row, column=4).value
    while matric:
        students[matric]={"firstname":firstname, "lastname":lastname, "route":route}
        row +=1
        matric = sheet.cell(row=row, column=1).value
        lastname = sheet.cell(row=row, column=2).value
        firstname = sheet.cell(row=row, column=3).value
        route = sheet.cell(row=row, column=4).value
    return students