# -*- coding: utf-8 -*-
"""
Created on Thu Jun 16 10:01:05 2022

@author: DMAMartin
"""
import os
import openpyxl


def findMRS(path, mod):
    module = mod.split()[0]
    try:
        files = [m for m in os.listdir(os.path.join(path, mod,'Assessment')) if m.startswith(module) and len(m)==25]
    except:
        return None
    if len(files)==1:
        return os.path.join(path, mod, "Assessment",files[0])
    else:
        print(module,'multiple files found: ', files)
        
def findMRS2023(path, mod):
    module = mod.split()[0]
    try:
        files = [m for m in os.listdir(os.path.join(path, mod)) if m.startswith(module) and len(m)==25]
    except:
        return None
    if len(files)==1:
        return os.path.join(path, mod,files[0])
    else:
        print(module,'multiple files found: ', files)

def find_modules2023(path):  
    
    mods = [f for f in os.listdir(path) if f[:2]=='BS']
    print('found ', mods)
    mrs = dict([(mod, findMRS2023(path, mod)) for mod in mods if findMRS2023(path,mod)])
    
        #mods = [m for m in mods if m in mrs]
    return mrs
        
def find_modules(path):  
    TA1 = os.path.join(path,"Level 1")
    TA2 = os.path.join(path,"Level 2")
    L3Bio = os.path.join(path,"Level 3", "Biological")
    L3Bms = os.path.join(path,"Level 3", "Biomedical")
    L4Bio = os.path.join(path,"Level 4", "Biological Stream")
    L4Bms = os.path.join(path,"Level 4", "Biomedical Stream")
    
    mods = [f for f in os.listdir(TA1) if f[:2]=='BS']
    mrs = dict([(mod, findMRS(TA1, mod)) for mod in mods if findMRS(TA1,mod)])
    for route in (TA2, L3Bio, L3Bms, L4Bio,L4Bms):
        mods2 = [f for f in os.listdir(route) if f[:2]=='BS']
        for m in mods2:
            if  findMRS(route,m):
                mrs[m] = findMRS(route,m) 
                mods.append(m)
        #mods = [m for m in mods if m in mrs]
    return mrs


def extractStudents(file):
    workbook=openpyxl.load_workbook(filename=file, data_only=True)
    sheet = workbook['OVERALL Results']
    row = 5
    students={}
    matric = sheet.cell(row=row, column=1).value
    lastname = sheet.cell(row=row, column=2).value
    firstname = sheet.cell(row=row, column=3).value
    route = sheet.cell(row=row, column=4).value
    while matric:
        students[matric]={"firstname":firstname, "lastname":lastname, "route":route}
        row +=1
        matric = sheet.cell(row=row, column=1).value
        lastname = sheet.cell(row=row, column=2).value
        firstname = sheet.cell(row=row, column=3).value
        route = sheet.cell(row=row, column=4).value
    return students
    