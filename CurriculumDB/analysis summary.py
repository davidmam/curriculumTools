# -*- coding: utf-8 -*-
"""
Created on Thu Jun 16 10:48:00 2022

@author: DMAMartin
"""

from tkinter import Tk
from tkinter.filedialog import askdirectory
import os
import openpyxl
from MRS import find_modules

def extractanalysis(file):
    workbook=openpyxl.load_workbook(filename=file, data_only=True)
    sheet = workbook['ANALYSIS']
    
    grades={}
    for row in range(5,32):
        count = sheet.cell(row=row, column=11).value
        grade = sheet.cell(row=row, column=10).value
    
        grades[grade]=count
    for row in range(5,11):
        count = sheet.cell(row=row, column=16).value
        grade = sheet.cell(row=row, column=15).value
        
        grades[grade]=count
    for row in range(5,11):
        count = sheet.cell(row=row, column=18).value
        grade = sheet.cell(row=row, column=19).value
        
        grades[grade]=count
    return grades

def main():
    root = Tk()
    
    TAdir = askdirectory(title="TA folder")
    root.destroy()
    mrs = find_modules(TAdir)
    mods = [x for x in mrs if x[2] in '12']
    gradedist = {}
    for m in mods:
        code = m.split()[0]
        gradedist[code]=extractanalysis(mrs[m])
    ofh = open('modulesummary.txt','w')
    rows='''A1
A2
A3
A4
A5
B1
B2
B3
C1
C2
C3
D1
D2
D3
M1
M2
M3
CF
BF
AB
QF
MC
CA
ST
WD
DC
NM
Number starting
Number withdrawing (WD,DC and ST)
Number passing diet 1
Number passing both diets combined
Module pass rate (inc WD,DC and ST)
Module pass rate (exc WD,DC and ST)
A
B
C
D
M
CF/BF'''.split('\n')
    modules = sorted([x.split()[0] for x in mods])
    print("\t",'\t'.join(modules), sep='', file=ofh)
    for r in rows:
        print(r,end='\t',file=ofh)
        for m in modules:
            print(gradedist[m][r],end='\t',file=ofh)
        print(file=ofh)
    ofh.close()

        
    
    
    
    
    