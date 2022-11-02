# -*- coding: utf-8 -*-
"""
Created on Mon Sep 26 09:41:18 2022
Load module information into the redis DB
@author: dmamartin

"""

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import redis
import os
from redisgraph import Node, Edge, Graph, Path

r = redis.Redis(host='redis-19330.c226.eu-west-1-3.ec2.cloud.redislabs.com', port=19330, username='cdev', password=open('redispassword.txt').read())

redis_graph = Graph('curriculum', r)

wb=openpyxl.open("zz MASTER MODULE MANAGERS LIST Office 2022-23.xlsx")
sheet=wb['Master']
nodes=[]
headers = [ str(x.value).strip().replace(' ','_').replace('C/','') for x in sheet[3]]
nodedata = []    

for row in range(4,90):
    params = {'Academic_Year':'2122'}
    for c in range(12):
        val = sheet.cell(row=row, column=c+1).value
        if val:
            params[headers[c]] = val
    print(params)
    params['name'] = params['Module_code']+" "+params['Module_Name']
    nodedata.append(params)
    nodes.append(Node(label='Module', properties=params))

for n in nodes:
    redis_graph.add_node(n)    


graphnodes={}

for n in redis_graph.nodes:
    graphnodes[redis_graph.nodes[n].properties['Module_code']] = redis_graph.nodes[n]
coremodules={}
for m in nodedata:
    if m.get('Optonal_Modules','') == 'C':
        if m['Level'] not in coremodules:
            coremodules[m['Level']] ={}
        if m['Semester'] not in coremodules[m['Level']]:
            coremodules[m['Level']][m['Semester']]=[]
        coremodules[m['Level']][m['Semester']].append(m)

for l in range(1,5):
    level1 = int((l)/2)+1
    level2 = int((l+1)/2)
    semester1 = l%2 +1
    semester2 = (l+1)%2+1
    for m in coremodules[level1][semester1]:
        for n in coremodules[level2][semester2]:
            redis_graph.add_edge(Edge(graphnodes[m['Module_code']], 'HAS_PREREQUISITE', graphnodes[n['Module_code']]))
    print(F'L{level1} S{semester1} L{level2} S{semester2}')
    
routes = {'BIMS': ["BMS","Biomedical Sciences"],
 'NEUR': ["BMS","Neuroscience"],
 'PHAR': ["BMS","Pharmacology"],
 'PHSC': ["BMS","Physiological Sciences"],
 'BIOLOGSCI': ["BIO","Biological Sciences"],
 'BIOC': ["BIO","Biochemistry"],
 'BSBI' : ["BIO","Biological Sciences (Bioinformatics)"],
 'BSPS': ["BIO","Biological Sciences (Plant Sciences)"],
 'MBIO': ["BIO","Microbiology"],
 'MOLG': ["BIO","Molecular Genetics"],
 'MOLB':["BIO","Molecular Biology"],
 'BCDD' : ["BIO","Biological Chemistry and Drug Discovery"]}

prognodes={}

for d in routes:
    prognodes[d] = Node(label='Programme',properties={'name': routes[d][1], 'stream':routes[d][0], 'code':d})
    redis_graph.add_node(prognodes[d])
#redis_graph.commit()

wb=openpyxl.open("Tabular programme data - AC 28092022.xlsx")

redis_graph = Graph('curriculum', r)
result=redis_graph.query('''MATCH (m:Module) return m''')
modules = {}
for m in result.result_set:  
    modules[m[0].properties['Module_code']]=m[0]
result=redis_graph.query('''MATCH (p:Programme) return p''')
programmes = {}
for p in result.result_set:  
    programmes[p[0].properties['name']]=m[0]
coreopt = {'C': 'REQUIRED', 'O': 'OPTIONAL'}    
sheet = wb['Program spec']
row = 2
stream = sheet.cell(row=row, column=1).value
while stream:
    #print([x.value for x in sheet[row]])
    try:
        prog=sheet.cell(row=row, column=2).value
        if prog:
            prog = prog.split()[-1][1:-1]
        
        if stream in programmes:
            module=sheet.cell(row=row, column=3).value[:7]
            core = sheet.cell(row=row, column=4).value
            if core not in coreopt:
                print(core, row)
                row=row+1
                stream = sheet.cell(row=row, column=1).value
                continue
            if module in modules:
                params=  {'module':module, 'name':prog,'coreopts':coreopt[core] }
                #print(params)
                if prog:
                    if prog in routes:
                        #pass
                        redis_graph.query("""MATCH (p:Programme {code:$name}) MATCH (m:Module {Module_code:$module})  MERGE (p)-[co:MODULE_TAKING {optional:$coreopts}]->(m)""", params)
                    else:
                        print (F'Unknown route {prog}')
                else:
                    for route in routes:
                        params = {'module':module, 'name':route,'coreopts':coreopt[core] }
                        #print(params)
                        redis_graph.query("""MATCH (p:Programme {code:$name}) MATCH (m:Module {Module_code:$module})  MERGE (p)-[co:MODULE_TAKING {optional:$coreopts}]->(m)""", params)
        else:
            print(F'programme {stream} not found')
        row = row+1
        stream = sheet.cell(row=row, column=1).value
    except Exception as e:
        print(row,module, prog, core, coreopt[core], e)
        row=row+1
            
milos ={}
milocount = 0

sheet = wb['MILO']

row = 2
module = sheet.cell(row=row, column=1).value
while module:
    module = module[:7]
    milo = sheet.cell(row=row, column=2).value
    learning_type = sheet.cell(row=row, column=3).value
    if milo not in milos:
        milocount+=1
        milos[milo]={'milo_id':milocount, 'modules':[], 'type':learning_type}
    milos[milo]['modules'].append(module)
    row = row+1
    module = sheet.cell(row=row, column=1).value
for m in  milos:
    try:
        redis_graph.query("MERGE (milo:MILO {milo_id:$mid, objective:$objective, type:$type})", {'mid':milos[m]['milo_id'], 'objective':m, 'type':milos[m]['type']})
    except:
        print("MERGE (milo:MILO {milo_id:$mid, objective:$objective, type:$type})", {'mid':milos[m]['milo_id'], 'objective':m, 'type':milos[m]['type']})   
    for mod in milos[m]['modules']:
        try:
            redis_graph.query("MATCH (milo:MILO {milo_id:$mid} ) MATCH (m:Module {Module_code:$mod}) MERGE (m) -[:HAS_OBJECTIVE]->(milo)", {'mid':milos[m]['milo_id'], 'mod':mod})
        except:
            print("MATCH (milo:MILO {milo_id:$mid} ) MATCH (m:Module {Module_code:$mod}) MERGE (m) -[:HAS_OBJECTIVE]->(milo)", {'mid':milos[m]['milo_id'], 'mod':mod})   
pilos ={}
pilocount = 0

sheet = wb['PILO']

row = 2
prog = sheet.cell(row=row, column=1).value
while prog:
    
    pilo = sheet.cell(row=row, column=2).value.strip()
    learning_type = sheet.cell(row=row, column=3).value
    if pilo not in pilos:
        pilocount+=1
        pilos[pilo]={'pilo_id':pilocount, 'programmes':[], 'type':learning_type}
    pilos[pilo]['programmes'].append(prog)
    row = row+1
    prog = sheet.cell(row=row, column=1).value
for p in  pilos:
    redis_graph.query("MERGE (pilo:PILO {pilo_id:$pid, objective:$objective, type:$type})", {'pid':pilos[p]['pilo_id'], 'objective':p, 'type':pilos[p]['type']})
    for pro in pilos[p]['programmes']:
        redis_graph.query("MATCH (pilo:PILO {pilo_id:$pid} ) MATCH (p:Programme {name:$prog}) MERGE (p) -[:HAS_OBJECTIVE]->(pilo)", {'pid':pilos[p]['pilo_id'], 'prog':pro})
        if pro == "Biological Sciences":
            for x in ("BIOC", "BSBI","BSPS","MBIO","MOLG","MOLB","BCDD"):
                xpro = routes[x][1]
                redis_graph.query("MATCH (pilo:PILO {pilo_id:$pid} ) MATCH (p:Programme {name:$prog}) MERGE (p) -[:HAS_OBJECTIVE]->(pilo)", {'pid':pilos[p]['pilo_id'], 'prog':xpro})
                
wb=openpyxl.open("QAA mapping biosciences.xlsx")
sheet = wb['Sheet1']
row = 2
subject = sheet.cell(row=row, column=1).value
while subject:
    data = [str(x.value).strip() for x in sheet[row]]
    params =  {'subject': data[0],
        'section': data[1],
        'subsection': data[2],
        'descriptor': data[3],
        'section_text': data[4],
        'subsection_text': data[5],
        'descriptor_text': data[6]       
        }
    redis_graph.query("MERGE (:QAA_benchmark {subject:$subject, section:$section, subsection:$subsection, descriptor:$descriptor,section_text:$section_text, subsection_text:$subsection_text, descriptor_text:$descriptor_text}) ",params)
    row=row+1
    subject = sheet.cell(row=row, column=1).value
wb=openpyxl.open("QAA biomedical benchmarks.xlsx")
sheet = wb['Sheet1']
row = 2
subject = sheet.cell(row=row, column=1).value
while subject:
    data = [str(x.value).strip() for x in sheet[row]]
    params =  {'subject': data[0],
        'section': data[1],
        'subsection': data[2],
        'descriptor': data[3],
        'section_text': data[4],
        'subsection_text': data[5],
        'descriptor_text': data[6]       
        }
    redis_graph.query("MERGE (:QAA_benchmark {subject:$subject, section:$section, subsection:$subsection, descriptor:$descriptor,section_text:$section_text, subsection_text:$subsection_text, descriptor_text:$descriptor_text}) ",params)
    row=row+1
    subject = sheet.cell(row=row, column=1).value

mabfolder = r'S:\Lifesci\LifeSciOff\SLT\2122\SLSLT Teaching Admin\Academic Standards\MAB Structures APP 2122\MABS'

mablist = [x for x in os.listdir(mabfolder) if x.startswith('BS')]

wb=openpyxl.open(os.path.join(mabfolder,mablist[0]))
sheet =wb['MAB']
type(sheet[12][8].value)

asstypes={}
for x in mablist:
    wb=openpyxl.open(os.path.join(mabfolder,x),data_only=True)
    sheet =wb['MAB']
    module = sheet[5][10].value
    credit = sheet[3][13].value
    row=12
    assessment =sheet[row][9].value
    while assessment:
        rowdata = [str(v.value) for v in sheet[row]]
        sequence = None
        try:
            sequence = int(rowdata[8])
        except:
            print('formative',rowdata, [v.value for v in sheet[row]])
        
        if sequence:
            params={'seq':rowdata[8],
                    'name':rowdata[9],
                    'ast':rowdata[10],
                    'type':rowdata[11],
                    'weight':rowdata[12],
                    'credits':str(int(rowdata[12])*credit/100),
                    'module':module,
                    }
            asstypes[rowdata[11]]=asstypes.get(rowdata[11], 0)+1
            redis_graph.query("MATCH (b:Module {Module_code:$module}) merge (a:Assessment {sequence:$seq, name:$name, AST_code:$ast, Assessment_type:$type, weight_percent:$weight, credits:$credits,module:$module})  MERGE (b)-[:HAS_ASSESSMENT]->(a)", params)
        else:
            params={'seq':rowdata[8],
                    'name':rowdata[9],
                    'ast':rowdata[10],
                    'type':rowdata[11],
                    'module':module,
                    }
            redis_graph.query("MATCH (b:Module {Module_code:$module}) merge (a:TeachingActivity {activity_type:'Formative Assessment', name:$name, AST_code:$ast, Assessment_type:$type,module:$module})  MERGE (b)-[:HAS_ACTIVITY]->(a)", params)
            #print("MERGE (a:Assessment {sequence:$seq, name:$name, AST_code:$ast, Assessmnet_type:$type, weight_percent:$weight, credits:$credits,module:$module}) MERGE (b) -[:HAS_ASSESSMENT]->(a)", params)
        row +=1
        assessment =sheet[row][9].value

timetables=r"C:\Users\dmamartin\OneDrive - University of Dundee\2223 TIMETABLE PREPARATION\2122 Timetables"
        
def find_all_timetables(path):
    
    routes=[]
    for sem in os.listdir(path):
        print(sem)
        for level in os.listdir(os.path.join(path,sem)):
            print(level)
            for module in os.listdir(os.path.join(path,sem,level)):
                print(module)
                if module.startswith('BS') and module.endswith('2.xlsx'):
                    routes.append(os.path.join(path,sem,level,module))
    return routes

ttlist = find_all_timetables(timetables)

for t in ttlist:
    module = t.split('\\')[-1].split()[0]
    wb=openpyxl.open(t)
    sheet =wb.active
    for row in list(sheet.rows)[3:]:
        if not row[4].value:
            print('No event name', [x.value for x in row])
            break
        event=row[4].value.split(' /')[0]
        atype = row[5].value
        week = row[2].value
        duration=str(row[8].value)
        params={
                'name':F'{event} week {week}',
                'type':atype,
                'duration':duration,
                'module':module,
                }
        if event:
            #print(event, "MATCH (b:Module {Module_code:$module}) merge (a:TeachingActivity {activity_type:$type, name:$name, duration:$duration,module:$module})  MERGE (b)-[:HAS_ACTIVITY]->(a)", params)
            redis_graph.query("MATCH (b:Module {Module_code:$module}) merge (a:TeachingActivity {activity_type:$type, name:$name, duration:$duration,module:$module})  MERGE (b)-[:HAS_ACTIVITY]->(a)", params)
            
             
def make_unique():    
    unique = '''MATCH (p:Module)
    WITH p.id as id, collect(p) AS nodes 
    WHERE size(nodes) >  1
    UNWIND nodes[1..] AS node
    DELETE node'''    
    
    redis_graph.query(unique)

# Create mapping spreadsheets

# get PILO from DB
pilolist = []
result = redis_graph.query('Match (p:PILO) return p.pilo_id, p.objective',{})
for pilo in result.result_set:
    pilolist.append(F'{pilo[0]:02d} - {pilo[1]}')

#set some default fonts

headerfont = font = Font(name='Calibri',
                 size=14,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')
tableheaderfont = font = Font(name='Calibri',
                 size=12,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')

setwrap = Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=True,
                     indent=0)

box = Border(left=Side(border_style='thick',
                          color='FF000000'),
                 right=Side(border_style="thick",
                            color='FF000000'),
                 top=Side(border_style="thick",
                          color='FF000000'),
                 bottom=Side(border_style="thick",
                             color='FF000000'),
                 diagonal=Side(border_style=None,
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style=None,
                              color='FF000000'),
                 vertical=Side(border_style=None,
                               color='FF000000'),
                 horizontal=Side(border_style=None,
                                color='FF000000')
                )



wb = openpyxl.Workbook()
ws1=wb.active
ws1.title = "ProgrammeLO"
for p in sorted(pilolist):
    ws1.append([p])
    
piloform= F"{quote_sheetname(ws1.title)}!$A$1:$A${len(pilolist)}"
dv = DataValidation(type="list", formula1=piloform, allow_blank=True)
dv.prompt = 'Select the most appropriate Programme Learning Objective'

result = redis_graph.query('MATCH (m:Module) where m.Module_code is not null  return m.Module_code, m.Module_Name',{})
for m in sorted(result.result_set)[1:]:
    if not m[0]:
        continue    
    print(m[0],m[1])
    if m[0] not in wb: 
        wb.create_sheet(m[0])
    ws = wb[m[0]]
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 120
    
    ws['A1']=m[0]
    ws['A1'].font = headerfont
    ws['A2']=m[1]
    ws['A2'].font = headerfont
    ws['A4']="Module Learning Objective"
    ws['B4']="Programme Learning Objective"
    ws['A4'].font=tableheaderfont
    ws['B4'].font=tableheaderfont
    row =5
    ws.add_data_validation(dv)
    mres = redis_graph.query('MATCH (m:Module {Module_code:$code} ) -[]-> (n:MILO) return n.milo_id, n.objective',{'code':m[0]})
    for mip in mres.result_set:
        ws[F'A{row}'] = F'{mip[0]} - {mip[1]}'
        ws[F'A{row}'].alignment=setwrap
        ws[F'B{row}'].alignment=setwrap
        dv.add(ws[F'B{row}'])
        row =row+1


wb.save('LOmapping.xlsx')    

## import HAN data
wb=openpyxl.open('Template BOKS BML-juni2017 (1) EN.xlsx', data_only=True)
ws = wb['Template BOKS BML-juni2017 (1) ']

row = 2
data = [x.value for x in ws[row]]
try:
    while data[0]:
        row = row+1
        for d in range(8):
            if data[d] is None:
                data[d] =''
        redis_graph.query('create (:BOKS {subject:$section, subset:$subset, code:$code, LOdescription:$LO, year_1:$y1, year_2:$y2, year_34:$y34, years_taught:$yt})',
                          {'section':data[0],'subset':data[1],'code':data[2],'LO':data[3],'y1':data[4],'y2':data[5], 'y34':data[6],'yt':data[7]})
        data = [x.value for x in ws[row]]
except Exception as e:
    print(row, data, e)

# create module mapping to assessment spreadsheet

module_assessments={}
result= redis_graph.query("MATCH (m:Module) -[]-> (a:Assessment) return m.Module_code, a.sequence, a.name, a.Assessment_type, a.weight_percent")
for d in result.result_set :
    if d[0] not in module_assessments:
        module_assessments[d[0]]=[]
    module_assessments[d[0]].append(d)

moduleLO = {}
result= redis_graph.query("MATCH (m:Module) -[]-> (a:MILO) return m.Module_code, a.milo_id, a.objective")
for d in result.result_set :
    if d[0] not in moduleLO:
        moduleLO[d[0]]=[]
    moduleLO[d[0]].append(d)

wb = openpyxl.Workbook()
ws=wb.active
ws.title = "Assessment to LO mapping"

for p in sorted(moduleLO):
    if p not in module_assessments:
        continue
    if p not in wb: 
        wb.create_sheet(p)
        
    ws =  wb[p]
    miloform = F"$A$3:$A$4"
    ws['A3']='Assessed'
    ws['A4']='Not Assessed'
    ws.row_dimensions[3].hidden = True
    ws.row_dimensions[4].hidden = True
    dv = DataValidation(type="list", formula1=miloform, allow_blank=True)
    dv.prompt = ''
    ws.column_dimensions['A'].width = 120
    
    assess = sorted(module_assessments[p], key = lambda x: x[1])
    for c in range(len(module_assessments[p])):
        try:
            ws.column_dimensions['BCDEFGHIJ'[c]].width = 30
            ws[F'{"BCDEFGHI"[c]}5']='Seq 00'+str(assess[c][1])
            ws[F'{"BCDEFGHI"[c]}6']=str(assess[c][2])
            ws[F'{"BCDEFGHI"[c]}7']=str(assess[c][3])
            ws[F'{"BCDEFGHI"[c]}8']=str(assess[c][4])+'%'
        except Exception as e:
            print(p,c, module_assessments[p], e)
            break
    ws['A1']=p
    ws['A1'].font = headerfont
    ws['A2']='Module assessment to ILO mapping'
    ws['A5']="Seq"
    ws['A6']="Name"
    ws['A7']="Assessment type"
    ws['A8']="Weight percent"
    ws['A5'].font=tableheaderfont
    ws['A6'].font=tableheaderfont
    ws['A7'].font=tableheaderfont
    ws['A8'].font=tableheaderfont
    row =9
    
    ws.add_data_validation(dv)
    for mip in moduleLO[p]:
        ws[F'A{row}'] =F'{mip[1]} - {mip[2]}'
        ws[F'A{row}'].alignment=setwrap
        for c in range(len(module_assessments[p])):
            dv.add(ws[F'{"BCDEFGHI"[c]}{row}'])
            ws[F'{"BCDEFGHI"[c]}{row}'].border = box
        row =row+1
   
wb.save('AssessmentLOmapping.xlsx') 